import math
from zoneinfo import ZoneInfo
import pandas as pd
from typing import Any, Dict, List, Optional
import time
import requests
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Tuple, Optional
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

TZ_AR = ZoneInfo("America/Argentina/Buenos_Aires")

# =========================
# CONSTANTES
# =========================
REDASH_QUERY_ID = 20036
REDASH_API_KEY = "tIuuysHOXHdN7WArCwGAFrJ9byyZrfBuY4svmMtS"

INCIDENCES_MAP = {
    "ABSENT": "Ausencia sin aviso",
    "LATE": "Tardanza",
    "UNDERWORKED": "Trabajo insuficiente",
    "LOCATION_INCIDENCE": "Fuera de ubicación"
}

WEEKDAY_ES_MAP = {
    0: "Lunes",
    1: "Martes",
    2: "Miércoles",
    3: "Jueves",
    4: "Viernes",
    5: "Sábado",
    6: "Domingo",
}

# =========================
# FECHAS / HORAS
# =========================

def iso_to_dt(value, tz=TZ_AR):
    if not value:
        return pd.NaT
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).astimezone(tz)
    except Exception:
        return pd.NaT


def floor_minute(dt):
    if pd.isna(dt):
        return dt
    return dt.replace(second=0, microsecond=0)


def weekday_es(date_str: str) -> str:
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d")
        return WEEKDAY_ES_MAP[d.weekday()]
    except Exception:
        return ""


def fmt_range(start, end):
    if pd.isna(start) or pd.isna(end):
        return ""
    return f"{start.strftime('%H:%M')} - {end.strftime('%H:%M')}"


def calc_delta_hours(real, sched, tolerance_seconds=0):
    if pd.isna(real) or pd.isna(sched):
        return 0.0
    delta = (real - sched).total_seconds() - tolerance_seconds
    return round(delta / 3600, 2) if delta > 0 else 0.0

def calc_early_arrival_hours(real, sched):
    if pd.isna(real) or pd.isna(sched):
        return 0.0

    delta = (sched - real).total_seconds()
    return round(delta / 3600, 2) if delta > 0 else 0.0


# =========================
# CATEGORÍAS / HORAS
# =========================

def split_categorized_hours(categorized_hours, categorias_validas):
    """
    Devuelve dict:
    HORAS_<CATEGORIA> = horas
    """
    out = {}
    for c in categorias_validas:
        out[f"HORAS_{c}"] = 0.0

    for ch in categorized_hours or []:
        name = (ch.get("category", {}).get("name") or "").upper().strip()
        if name in categorias_validas:
            out[f"HORAS_{name}"] += float(ch.get("hours") or 0)

    return {k: round(v, 2) for k, v in out.items()}


# =========================
# INCIDENCIAS
# =========================
def build_observaciones(it: dict) -> str:
    obs = []

    # =========================
    # Feriados
    # =========================
    holidays = it.get("holidays") or []
    if isinstance(holidays, list) and holidays:
        names = []
        for h in holidays:
            if isinstance(h, dict):
                n = (h.get("name") or "").strip()
                if n:
                    names.append(n)

        if names:
            # dedupe simple (orden no importa tanto en feriados)
            obs.append("Feriado: " + " | ".join(sorted(set(names))))

    # =========================
    # Incidencias
    # - puede venir como ["ABSENT", ...]
    # - o como [{"name": "..."}]
    # =========================
    incidences = it.get("incidences") or []
    if isinstance(incidences, list) and incidences:
        names = []
        for inc in incidences:
            if isinstance(inc, str):
                n = inc.strip()
                if n:
                    key = n.upper().strip()
                    label = INCIDENCES_MAP.get(key, n)  # 👈 traducción
                    names.append(label)
            elif isinstance(inc, dict):
                n = (inc.get("name") or inc.get("type") or inc.get("code") or "").strip()
                if n:
                    names.append(n)

        if names:
            # dedupe manteniendo orden
            seen = set()
            names_unique = []
            for n in names:
                if n not in seen:
                    seen.add(n)
                    names_unique.append(n)

            obs.append("Incidencia: " + " | ".join(names_unique))

    # =========================
    # Time off requests (Licencias)
    # =========================
    tors = it.get("timeOffRequests") or []
    if isinstance(tors, list) and tors:
        names = []
        for tor in tors:
            if isinstance(tor, dict):
                n = (tor.get("name") or "").strip()
                if n:
                    names.append(n)

        if names:
            # dedupe manteniendo orden
            seen = set()
            names_unique = []
            for n in names:
                if n not in seen:
                    seen.add(n)
                    names_unique.append(n)

            obs.append("Licencia: " + " | ".join(names_unique))

    return " | ".join(obs)


def clasificar_empleado_por_scheduled_max(df, col_sched="SCHEDULED_HOURS"):
    """
    FULL-TIME  : max scheduled >= 8
    PART-TIME  : max scheduled < 8
    """
    res = {}
    grouped = df.groupby("ID")[col_sched].max()

    for emp, max_h in grouped.items():
        try:
            max_h = float(max_h or 0)
        except Exception:
            max_h = 0

        if max_h >= 8:
            res[emp] = "FULL-TIME"
        else:
            res[emp] = "PART-TIME"

    return res


# =========================
# NOCTURNIDAD
# =========================

def nocturnidad_es_100(row):
    """
    Regla:
    - Domingo
    - Feriado
    - No laborable
    - Sábado (fallback conservador)
    """
    weekday = row.get("_weekday_api", "")
    if weekday in ("SUNDAY",):
        return True

    if row.get("_hasHoliday_api", False):
        return True

    if row.get("_isWorkday_api") is False:
        return True

    if weekday == "SATURDAY":
        return True

    return False


# =========================
# STRINGS / NOMBRES
# =========================

def split_apellido_nombre(value):
    if not value or "," not in value:
        return "", ""
    apellido, nombre = value.split(",", 1)
    return apellido.strip(), nombre.strip()


# =========================
# EXPORTACIÓN EXCEL
# =========================

def horas_para_excel(value, usar_decimal=True):
    """
    - 0 => celda vacía
    - decimal => float
    - hh:mm => fracción de día
    """
    try:
        v = float(value)
    except Exception:
        return ""

    if v == 0:
        return ""

    if usar_decimal:
        return round(v, 2)

    return v / 24.0



def export_detalle_diario_excel(
    df_export: pd.DataFrame,
    out: str,
    START_DATE: str,
    END_DATE: str,
    generated_at: str,
    EXPORTAR_DECIMAL: bool,
    COLS_HORAS_DETALLE: list,
):
    with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
        workbook = writer.book

        fmt_title = workbook.add_format({"bold": True, "font_size": 14})
        fmt_sub   = workbook.add_format({"font_size": 11})
        fmt_wrap  = workbook.add_format({"text_wrap": True})
        fmt_hhmm  = workbook.add_format({"num_format": "[h]:mm"}) if not EXPORTAR_DECIMAL else None
        fmt_dec   = workbook.add_format({"num_format": "0.00"})   if EXPORTAR_DECIMAL else None

        # ----- Detalle diario -----
        startrow = 4

        # Ordenar
        df_export = df_export.sort_values(by=["ID", "Fecha"], ascending=[True, True]).reset_index(drop=True)

        # Escribir dataframe
        df_export.to_excel(writer, index=False, sheet_name="Detalle diario", startrow=startrow)
        ws1 = writer.sheets["Detalle diario"]

        # Encabezado
        ws1.write("A1", "DETALLE DIARIO DE ASISTENCIA", fmt_title)
        ws1.write("A2", f"Período: {START_DATE} al {END_DATE}", fmt_sub)
        ws1.write("A3", f"Generado: {generated_at}", fmt_sub)

        # Anchos + formatos
        for idx, col in enumerate(df_export.columns):
            if col == "Observaciones":
                ws1.set_column(idx, idx, 45, fmt_wrap)
            elif col in COLS_HORAS_DETALLE:
                ws1.set_column(idx, idx, 22, fmt_dec if EXPORTAR_DECIMAL else fmt_hhmm)
            else:
                ws1.set_column(idx, idx, 26)

    print("✅ Excel generado:", out)


def aplicar_regla_extra_50(horas_extra):
    """
    Regla:
    - <= 0.5  -> 0
    - > 0.5   -> baja a escalones de 0.5
      0.6..1.0 -> 0.5
      1.1..1.5 -> 1.0
      1.6..2.0 -> 1.5
      etc.
    """
    try:
        h = float(horas_extra)
    except Exception:
        return 0.0

    if h <= 0.5:
        return 0.0

    # ejemplo: 1.1 -> int((1.1 - 0.5) / 0.5) = 1  -> 0.5
    escalones = int((h - 0.5) / 0.5)
    return round(escalones * 0.5, 2)


def decimal_to_hhmm(x):
    if pd.isna(x) or x == "":
        return ""
    minutes = int(round(float(x) * 60))
    h = minutes // 60
    m = minutes % 60
    return f"{h:02d}:{m:02d}"


import time
import requests
import pandas as pd
from typing import Dict, List, Tuple, Any

def redash_fetch_rows(
                                            query_id: int,
                                            api_key: str,
                                            base_url: str = "https://redash.humand.co",
                                            session: Optional[requests.Session] = None,
                                            do_refresh: bool = True,
                                            refresh_wait_s: float = 2.0,
                                            max_retries: int = 3,
                                            timeout: int = 30
                                        ) -> List[Dict[str, Any]]:

    """
    Trae rows de un query de Redash (opcional refresh).
    - Optimizado: si refresh falla, igual intenta traer results.
    - Reintentos simples.
    """
    sess = session or requests.Session()

    headers = {
        "Authorization": api_key,   # en tu Redash funciona así
        "User-Agent": "Python requests",
        "Content-Type": "application/json",
    }

    refresh_url = f"{base_url}/api/queries/{query_id}/refresh"
    results_url = f"{base_url}/api/queries/{query_id}/results"

    last_exc = None
    
    if do_refresh:
        for _ in range(max_retries):
            try:
                r = sess.post(refresh_url, headers=headers, timeout=timeout)
                # 200 OK o 202 Accepted suelen ser válidos
                if r.status_code in (200, 202):
                    break
            except Exception as e:
                last_exc = e
            time.sleep(0.6)

        # pequeña espera para que materialice
        time.sleep(refresh_wait_s)
    
    for _ in range(max_retries):
        try:
            r = sess.get(results_url, headers=headers, timeout=timeout)
            print(r.status_code)
            r.raise_for_status()
            data = r.json()
            rows = data.get("query_result", {}).get("data", {}).get("rows", [])
            return rows or []
        except Exception as e:
            last_exc = e
            time.sleep(0.8)

    # fallback duro (results.json?api_key=...)
    try:
        fallback = f"{base_url}/api/queries/{query_id}/results.json?api_key={api_key}"
        r = sess.get(fallback, timeout=timeout)
        r.raise_for_status()
        data = r.json()
        return data.get("query_result", {}).get("data", {}).get("rows", []) or []
    except Exception:
        if last_exc:
            raise last_exc
        return []

def build_permissions_index(rows: List[Dict[str, Any]]) -> Dict[Tuple[str, str], str]:
    """
    Indexa permisos por (employeeInternalId, 'YYYY-MM-DD') -> texto.
    ✅ Soporta rango inclusive: dd/mm -> dd2/mm2
    """

    def z2(x) -> str:
        try:
            return f"{int(str(x).strip()):02d}"
        except Exception:
            return ""

    def z4(x) -> str:
        try:
            return f"{int(str(x).strip()):04d}"
        except Exception:
            return ""

    def mk_date(yyyy: str, mm: str, dd: str) -> Optional[date]:
        if not (yyyy and mm and dd):
            return None
        try:
            return date(int(yyyy), int(mm), int(dd))
        except Exception:
            return None

    def parse_range_from_row(r: Dict[str, Any]) -> Optional[Tuple[date, date]]:
        # Año base (si no viene, no podemos armar fecha con seguridad)
        yyyy = z4(r.get("Año") if r.get("Año") is not None else r.get("yyyy"))
        mm1  = z2(r.get("mm"))
        dd1  = z2(r.get("dd"))

        start = mk_date(yyyy, mm1, dd1)
        if start is None:
            # fallback: si "Fecha" viene como YYYY-MM-DD, úsala como día único
            f = r.get("Fecha")
            if isinstance(f, str) and re.match(r"^\d{4}-\d{2}-\d{2}$", f.strip()):
                try:
                    d = datetime.strptime(f.strip(), "%Y-%m-%d").date()
                    return (d, d)
                except Exception:
                    return None
            return None

        # fin: dd2/mm2 (si no hay, es día único)
        mm2 = z2(r.get("mm2") or r.get("mes2"))
        dd2 = z2(r.get("dd2") or r.get("dia2"))

        if not (mm2 and dd2):
            return (start, start)

        end = mk_date(yyyy, mm2, dd2)
        if end is None:
            return (start, start)

        # Si el rango vino invertido, normalizamos
        if end < start:
            start, end = end, start

        return (start, end)

    def parse_time_from_row(r: Dict[str, Any]) -> str:
        hh = z2(r.get("Hora"))
        mi = z2(r.get("Minutos"))
        if hh and mi:
            return f"{hh}:{mi}"
        return ""

    idx: Dict[Tuple[str, str], List[str]] = {}

    for r in rows or []:
        emp = r.get("employeeInternalId")
        if emp is None:
            continue
        emp = str(emp).strip()
        if not emp:
            continue

        rng = parse_range_from_row(r)
        if rng is None:
            continue

        start, end = rng

        hhmm = parse_time_from_row(r)
        incidencia = r.get("incidencia")
        incidencia = "" if incidencia is None else str(incidencia).strip()
        motivo =  r.get("motivo")
        motivo = "" if motivo is None else str(motivo).strip()
        tipo = r.get("TipoSolicitud")
        tipo = "" if tipo is None else str(tipo).strip()


        # armo “detalles” (TipoSolicitud + Incidencia)
        detalles = " - ".join([x for x in [tipo, incidencia, motivo] if x])

        # texto final
        if hhmm and detalles:
            texto = f"{hhmm} ({detalles})"
        elif hhmm:
            texto = hhmm
        elif detalles:
            texto = detalles
        else:
            texto = "Permiso"

        # ✅ Expandir rango INCLUSIVE (incluye el último día)
        d = start
        while d <= end:
            key = (emp, d.isoformat())
            idx.setdefault(key, []).append(texto)
            d += timedelta(days=1)

    # Flatten: si hay varios permisos el mismo día, concatenar
    out: Dict[Tuple[str, str], str] = {}
    for k, arr in idx.items():
        # si querés mantener orden “original”, sacá sorted()
        out[k] = " | ".join(sorted(arr))

    return out

def apply_permissions_column(df_export: pd.DataFrame,
                             permissions_index: Dict[Tuple[str, str], str],
                             emp_col: str = "ID",
                             date_col: str = "Fecha",
                             out_col: str = "Permisos Pedidos Aprobados") -> pd.DataFrame:
    """
    Crea/llena columna de permisos haciendo match por (ID, Fecha).
    - Fecha debe ser 'YYYY-MM-DD' (en tu df_export lo es).
    """
    if out_col not in df_export.columns:
        df_export[out_col] = ""

    # por si vienen como int
    emp_series = df_export[emp_col].astype(str).str.strip()
    date_series = df_export[date_col].astype(str).str.slice(0, 10)

    keys = list(zip(emp_series, date_series))
    df_export[out_col] = [permissions_index.get(k, "") for k in keys]

    return df_export

def flags_incidencias_y_eventos(
    entries=None,
    slots=None,
    incidences=None,
    time_off_requests=None,
    holidays=None
) -> dict:
    """
    Devuelve flags 'Si' / 'No' para columnas:
    - Ausencia
    - Tardanza
    - Retiro anticipado
    - Trabajo Insuficiente
    - Es Feriado
    - Licencia
    """

    incidences = incidences or []
    time_off_requests = time_off_requests or []
    holidays = holidays or []

    # Normalizar incidencias a set de claves UPPER
    inc_keys = set()

    for inc in incidences:
        if isinstance(inc, str):
            key = inc.strip().upper()
        elif isinstance(inc, dict):
            key = (inc.get("name") or inc.get("type") or inc.get("code") or "").strip().upper()
        else:
            continue

        if key:
            inc_keys.add(key)

    return {
        "Ausencia": "Si" if "ABSENT" in inc_keys else "No",
        "Tardanza": "Si" if "LATE" in inc_keys else "No",
        "Trabajo Insuficiente": "Si" if "UNDERWORKED" in inc_keys else "No",
        "Es Feriado": "Si" if bool(holidays) else "No",
        "Licencia": "Si" if bool(time_off_requests) else "No",
    }



def colorear_flags_excel(path_xlsx: str, sheet_name: str, flag_cols: list):
    wb = load_workbook(path_xlsx)
    ws = wb[sheet_name]

    fill_si = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # verde
    fill_no = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # rojo

    # detectar fila header buscando nombres exactos
    header_row = None
    col_index = {}  # nombre -> idx (1-based)

    for r in range(1, 50):  # busca header en primeras 50 filas
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if values and any(v in flag_cols for v in values):
            header_row = r
            for c, v in enumerate(values, start=1):
                if v in flag_cols:
                    col_index[v] = c
            if col_index:
                break

    if not header_row:
        wb.close()
        raise ValueError(f"No encontré header para columnas flags en hoja '{sheet_name}'")

    start_data_row = header_row + 1
    last_row = ws.max_row

    for col_name, cidx in col_index.items():
        for r in range(start_data_row, last_row + 1):
            cell = ws.cell(row=r, column=cidx)
            val = cell.value
            if val == "Si":
                cell.fill = fill_si
            elif val == "No":
                cell.fill = fill_no

    wb.save(path_xlsx)
    wb.close()
    
def redondear_extra_media_hora(x):
    # x en horas decimales (float)
    v = pd.to_numeric(x, errors="coerce")
    if pd.isna(v) or v < 0.5:
        return 0.0
    return math.floor(v / 0.5) * 0.5